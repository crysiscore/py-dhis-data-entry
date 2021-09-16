from encodings.utf_8 import encode
from selenium import webdriver
from selenium.webdriver.support.ui import Select
import yaml
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from helperfunctions import *
from openpyxl import load_workbook
import datetime
from config.extra_config import *
from os import chdir, system
import sys

# DHIS/Excell Mapping files
indicators_files =['tb_prev.yaml','tx_tb.yaml','tx_rtt.yaml','tx_new.yaml','tx_ml.yaml','tx_curr.yaml',
                    'tx_pvls.yaml','additional_data.yaml','ptv_cpn.yaml', 'retention_on_art_dsd.yaml',
                    'dsd_models.yaml','12_month_retention.yaml','ats_parte_a_emergency_ward.yaml','ats_parte_a_inpatient_services.yaml',
                    'ats_parte_a_other_pitc.yaml', 'ats_parte_a_pediatric_services.yaml','ats_parte_a_vct.yaml',
                    'ats_parte_b_hts_index.yaml']

tb_prev_file_full_path          = "mapping/" + indicators_files[0]
tx_tb_file_full_path            = "mapping/" + indicators_files[1]
tx_rtt_file_full_path           = "mapping/" + indicators_files[2]
tx_new_file_full_path           = "mapping/" + indicators_files[3]
tx_ml_file_full_path            = "mapping/" + indicators_files[4]
tx_curr_file_full_path          = "mapping/" + indicators_files[5]
tx_pvls_file_full_path          = "mapping/" + indicators_files[6]
addit_data_file_full_path       = "mapping/" + indicators_files[7]
ptv_file_full_path_full_path    = "mapping/" + indicators_files[8]

## Retention  Form
ret_on_art_1_3_month_full_path  = "mapping/" + indicators_files[9]
ret_on_art_dsd_models_full_path = "mapping/" + indicators_files[10]
ret_on_art_12_month_full_path   = "mapping/" + indicators_files[11]

## ATS Form
## 
ats_parte_a_emergency_ward_full_path      = "mapping/" + indicators_files[12]
ats_parte_a_inpatient_services_full_path  = "mapping/" + indicators_files[13]
ats_parte_a_other_pitc_full_path          = "mapping/" + indicators_files[14]
ats_parte_a_pediatric_services_full_path  = "mapping/" + indicators_files[15]
ats_parte_a_vct_full_path                 = "mapping/" + indicators_files[16]
ats_parte_b_hts_index_full_path           = "mapping/" + indicators_files[17]


# Read dhis2 dhis_config.yaml  ( Uncoment the following lines if run on windows or linux)
# Windows
#dhis_config = open_config_file("C:\\py-dhis-data-entry\\config\\dhis_config.yaml")
# Linux

dhis_config = open_config_file("/home/agnaldo/Git/py-dhis-data-entry/config/dhis_config.yaml")

username = dhis_config['dhis2_username']
password = dhis_config['dhis2_password']
district = dhis_config['distrito']
us_name = dhis_config['unidade_sanitaria']
period = dhis_config['periodo']

if period[0:3] == "Mar":
     #coded_period= period.encode('latin-1')
     #decoded_period = coded_period.decode('utf-8')
     print(period)
     period = period.encode('ISO-8859-1')
     period = period.decode('utf-8')
     # Print('o Periodo e %s', temp.decode('utf-8') )
     print('o Periodo e %s', period )

form_name = dhis_config['formulario']
excell_location = dhis_config['excell_location']
dhis_url = dhis_config['dhis_url']
sheet_name = dhis_config['sheet_name']
override = dhis_config['override']
working_dir = dhis_config['working_dir']
os_type = dhis_config['os_type']
user_role = dhis_config['user_role']
param_check =True


# TODO check if dir exists
chdir(working_dir)
list_config = [username,password,district,us_name,period,form_name,excell_location,dhis_url,sheet_name,override]


# Open Log file
# (same directory) in append mode and 
log_file = open("logs.txt","a+") 
log_file.truncate(0)
log_file.seek(0)

# check dhis configuration parameters
for index, item  in enumerate(list_config):
     if item:
         if index ==2:
             if item not in dict_distritos_maputo:
                print("O distrito: %s ,  esta mal escrito no ficheiro de configuracoes dhis_config.yaml ou nao pertence a lista de distritos disponiveis no dhis." % item)
                log_file.write("O distrito: %s ,  esta mal escrito no ficheiro de configuracoes dhis_config.yaml ou nao pertence a lista de distritos disponiveis no dhis.\n" % item)
                log_file.write("Erro inesperado!! verifique  e corrige os erros  acima\n")
                sys.exit()
              
         elif index ==3:
              if item not in unidades_sanitarias:
                 print("A unidade sanitaria: %s ,  esta mal escrito no ficheiro de configuracoes dhis_config.yaml ou nao pertence a lista de US disponiveis no dhis." % item)
                 log_file.write("A unidade sanitaria:  %s ,   esta mal escrito no ficheiro de configuracoes dhis_config.yaml ou nao pertence a lista de US disponiveis no dhis.\n" % item)
                 log_file.write("Erro inesperado!! verifique  e corrige os erros  acima\n")
                 sys.exit()

         elif index ==4:
              
              if period not in periodos:
                  print("O periodo : %s ,  esta mal escrito no ficheiro de configuracoes dhis_config.yaml ou nao pertence a lista de periodos no dhis." % period)
                  log_file.write("O periodo : %s ,  esta mal escrito no ficheiro de configuracoes dhis_config.yaml ou nao pertence a lista de periodos no dhis.\n" % period)
                  log_file.write("Erro inesperado!! verifique  e corrige os erros  acima\n")
                  sys.exit()  
         elif index ==5:
              if item not in forms:
                 print("O formulario : %s ,  esta mal escrito no ficheiro de configuracoes dhis_config.yaml ou nao pertence a lista de formularios no dhis." % item)
                 log_file.write("O formulario : %s ,  esta mal escrito no ficheiro de configuracoes dhis_config.yaml ou nao pertence a lista de formularios no dhis.\n" % item)
                 log_file.write("Erro inesperado!! verifique  e corrige os erros  acima\n")
                 sys.exit()  
         elif index ==9:
              if item not in ['Sim', 'Nao']:
                 print("O paramentro override : %s ,  esta mal escrito no ficheiro de configuracoes dhis_config.yaml ." % item)
                 log_file.write("O paramentro override : %s ,  esta mal escrito no ficheiro de configuracoes dhis_config.yaml.\n" % item)
                 log_file.write("Erro inesperado!! verifique  e corrige os erros  acima\n")
                 sys.exit()  
       
     else:
         if index ==0:
              print("username is empty.")
              log_file.write("O parametro (username) nao pode ser vazio no ficheiro de configuracoes dhis_config.yaml, preencha e volte a executar\n")
              param_check = False
         elif index ==1:
              print("password is empty.")
              log_file.write("O parametro (password) nao pode ser vazio no ficheiro de configuracoes dhis_config.yaml, preencha e volte a executar\n")
              param_check = False
         elif index ==2:
              print("district is empty.")
              log_file.write("O parametro distrito (district) nao pode ser vazio no ficheiro de configuracoes dhis_config.yaml, preencha e volte a executar\n")
              param_check = False
         elif index ==3:
              print("us_name is empty.")
              log_file.write("O parametro nome da us (us_name) nao pode ser vazio no ficheiro de configuracoes dhis_config.yaml, preencha e volte a executar\n")
              param_check = False
         elif index ==4:
              print("periodo is empty.")
              log_file.write("O parametro (periodo) nao pode ser vazio no ficheiro de configuracoes dhis_config.yaml, preencha e volte a executar\n")
              param_check = False
         elif index ==5:
              print("form_name is empty.")
              log_file.write("O parametro nome do formulario (form_name) nao pode ser vazio no ficheiro de configuracoes dhis_config.yaml, preencha e volte a executar\n")
              param_check = False
         elif index ==6:
              print("excell_location is empty.")
              log_file.write("O parametro localizacao do excell (excell_location) nao pode ser vazio no ficheiro de configuracoes dhis_config.yaml, preencha e volte a executar\n")
              param_check = False
         elif index ==7:
              print("dhis_url is empty.")
              log_file.write("O parametro url do dhis (dhis_url) nao pode ser vazio no ficheiro de configuracoes dhis_config.yaml, preencha e volte a executar\n")
              param_check = False
         elif index ==8:
              print("sheet_name is empty.")
              log_file.write("O parametro nome ds planilha excell  (sheet_name) nao pode ser vazio no ficheiro de configuracoes dhis_config.yaml, preencha e volte a executar\n")
              param_check = False
         elif index ==9:
              print("override is empty.")
              log_file.write("O parametro override nao pode ser vazio no ficheiro de configuracoes dhis_config.yaml, preencha e volte a executar\n")
              param_check = False

if param_check: 

    try:
        # verifica se a us pertence ao devido distrito
        if check_us_in_district(us_name,district ):
             print("%s was found in %s ..." %(us_name,district))
             workbook = load_workbook('data/'+ excell_location)
               # grab the active worksheet
             if sheet_name in workbook.sheetnames:
                  active_sheet = workbook[sheet_name]
                  # Open chrome browser - Chrome options
                  #driver_loc = "C:/py-dhis-data-entry/drivers/chromedriver.exe"
                  if os_type == "linux":
                        driver_loc = "/usr/bin/chromedriver"
                  else:
                        driver_loc = "C:/py-dhis-data-entry/drivers/chromedriver.exe"
                       
                  #chrome_options = webdriver.ChromeOptions()

                  # default_directory  must be a configurable parameter, and thus should be written to a file
                  #chrome_options.add_argument(
                  #    "download.default_directory=/home/agnaldo/Git/py-dhis-data-entry/downloads")
                  chrome_browser = webdriver.Chrome(driver_loc) #Optional argument, if not specified will search path.
                  #
                  if   form_name == "PTV-Resumo Mensal de PTV" :
                       if check_ptv_template_integrity(active_sheet,log_file):
                            chrome_browser.get(dhis_url)
                            time.sleep(7)
                            chrome_browser.find_element_by_id("j_username").send_keys(username)
                            chrome_browser.find_element_by_id("j_password").send_keys(password)
                            chrome_browser.find_element_by_id("submit").click()

                            time.sleep(15)
                            chrome_browser.get(dhis_url + 'dhis-web-dataentry/index.action')
                            #tempo para a pagina terminar de carregar
                            time.sleep(10)
                            wait = WebDriverWait(chrome_browser, 15)
                            if user_role == "admin":
                                 xpath="//li[@id='orgUnitj9Inbtfw3Wu']/span/img[contains(@src, '/images/colapse.png')]"
                                 expand_root_tree =wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
                                 expand_root_tree.click()
                                 expand_province_tree('Cidade De Maputo',chrome_browser)
                                 expand_district_tree(district,chrome_browser)
                                 select_province(us_name, chrome_browser)
                                 #time.sleep(4)
                                 select_form(form_name,chrome_browser)
                                 #time.sleep(3)
                            else:
                                 expand_district_tree(district,chrome_browser)
                                 #time.sleep(4)
                                 select_province(us_name, chrome_browser)
                                 #time.sleep(4)
                                 select_form(form_name,chrome_browser)
                                 #time.sleep(3)
                            time.sleep(2)
                            now = datetime.datetime.now()
                            # Codificacao correcta de caracteres : problema com acentos
                            if str(now.year) in period:
                                 select_period(period,chrome_browser)
                            elif '2020' in period and str(now.year) == '2021':
                                  chrome_browser.find_element_by_id("prevButton").click() # prevButton for the previous year
                                  select_period(period,chrome_browser)
                            elif '2021' in period and str(now.year) == '2022':
                                  chrome_browser.find_element_by_id("prevButton").click() # prevButton for the previous year  
                                  select_period(period,chrome_browser)   
                            else:
                                 # This is less likely to happen
                                 chrome_browser.find_element_by_id("nextButton").click() # prevButton for the previous year
                                 select_period(period,chrome_browser)
                                 sys.exit('Nao foi possivel processar o periodo')

                            time.sleep(3)
                            fill_indicator_elements('ptv_cpn',ptv_file_full_path_full_path,active_sheet,log_file,chrome_browser,override)

                       else:
                             print('Verifique os problemas acima' )
                             log_file.write('Verifique os problemas acima' )
                             sys.exit('Verifique os problemas acima' )  
                       
                       #exccutar_validacao(chrome_browser)
                  elif form_name == "C&T_Resumo de Cuidados e Tratamento" :
                       if check_ct_template_integrity(active_sheet,log_file):
                            chrome_browser.get(dhis_url)
                            time.sleep(7)
                            chrome_browser.find_element_by_id("j_username").send_keys(username)
                            chrome_browser.find_element_by_id("j_password").send_keys(password)
                            chrome_browser.find_element_by_id("submit").click()

                            time.sleep(15)
                            chrome_browser.get(dhis_url + 'dhis-web-dataentry/index.action')
                            #tempo para a pagina terminar de carregar
                            time.sleep(10)
                            wait = WebDriverWait(chrome_browser, 15)
                            if user_role == "admin":
                                 xpath="//li[@id='orgUnitj9Inbtfw3Wu']/span/img[contains(@src, '/images/colapse.png')]"
                                 expand_root_tree =wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
                                 expand_root_tree.click()
                                 expand_province_tree('Cidade De Maputo',chrome_browser)
                                 expand_district_tree(district,chrome_browser)
                                 select_province(us_name, chrome_browser)
                                 #time.sleep(4)
                                 select_form(form_name,chrome_browser)
                                 #time.sleep(3)
                            else:
                                 expand_district_tree(district,chrome_browser)
                                 #time.sleep(4)
                                 select_province(us_name, chrome_browser)
                                 #time.sleep(4)
                                 select_form(form_name,chrome_browser)
                                 #time.sleep(3)
                            time.sleep(2)
                            now = datetime.datetime.now()
                            # Codificacao correcta de caracteres : problema com acentos
                            if str(now.year) in period:
                                 select_period(period,chrome_browser)
                            elif '2020' in period and str(now.year) == '2021':
                                  chrome_browser.find_element_by_id("prevButton").click() # prevButton for the previous year
                                  select_period(period,chrome_browser)
                            elif '2021' in period and str(now.year) == '2022':
                                  chrome_browser.find_element_by_id("prevButton").click() # prevButton for the previous year  
                                  select_period(period,chrome_browser)   
                            else:
                                 # This is less likely to happen
                                 chrome_browser.find_element_by_id("nextButton").click() # prevButton for the previous year
                                 select_period(period,chrome_browser)
                                 sys.exit('Nao foi possivel processar o periodo')
                                       
                            time.sleep(3)
                            #indicator_map_file,active_sheet,log_file,browser_webdriver)
                           
                            fill_indicator_elements('TB_PREV_NUMERATOR', tb_prev_file_full_path,active_sheet,log_file,chrome_browser,override)
                            fill_indicator_elements('TX_NEW', tx_new_file_full_path,active_sheet,log_file,chrome_browser,override)
                            fill_indicator_elements('TX_CURR', tx_curr_file_full_path,active_sheet,log_file,chrome_browser,override)
                            fill_indicator_elements('TX_RTT', tx_rtt_file_full_path,active_sheet,log_file,chrome_browser,override)
                            fill_indicator_elements('TX_ML', tx_ml_file_full_path,active_sheet,log_file,chrome_browser,override)
                            fill_indicator_elements('TX_TB', tx_tb_file_full_path,active_sheet,log_file,chrome_browser,override)
                            fill_indicator_elements('TX_PVLS', tx_pvls_file_full_path,active_sheet,log_file,chrome_browser,override) 
                            fill_indicator_elements('ADDITIONAL_DATA', addit_data_file_full_path,active_sheet,log_file,chrome_browser,override)

                       else:
                            print('Verifique os problemas acima' )
                            log_file.write('Verifique os problemas acima' )
                            sys.exit('Verifique os problemas acima' )  
                  elif form_name == "C&T_Relatorio de Retencao e Modelos Diferenciados de Saude":
                       if check_retencao_dsd_template_integrity(active_sheet, log_file):
                            chrome_browser.get(dhis_url)
                            time.sleep(7)
                            chrome_browser.find_element_by_id("j_username").send_keys(username)
                            chrome_browser.find_element_by_id("j_password").send_keys(password)
                            chrome_browser.find_element_by_id("submit").click()
                            time.sleep(15)
                            chrome_browser.get(dhis_url + 'dhis-web-dataentry/index.action')
                            #tempo para a pagina terminar de carregar
                            time.sleep(10)
                            wait = WebDriverWait(chrome_browser, 15)
                            if user_role == "admin":
                                 xpath="//li[@id='orgUnitj9Inbtfw3Wu']/span/img[contains(@src, '/images/colapse.png')]"
                                 expand_root_tree =wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
                                 expand_root_tree.click()
                                 expand_province_tree('Cidade De Maputo',chrome_browser)
                                 expand_district_tree(district,chrome_browser)
                                 select_province(us_name, chrome_browser)
                                 #time.sleep(4)
                                 select_form(form_name,chrome_browser)
                                 #time.sleep(3)
                            else:
                                 expand_district_tree(district,chrome_browser)
                                 #time.sleep(4)
                                 select_province(us_name, chrome_browser)
                                 #time.sleep(4)
                                 select_form(form_name,chrome_browser)
                                 #time.sleep(3)
                            time.sleep(2)
                            now = datetime.datetime.now()
                            # Codificacao correcta de caracteres : problema com acentos
                            if str(now.year) in period:
                                  select_period(period,chrome_browser)
                            elif '2020' in period and str(now.year) == '2021':
                                  chrome_browser.find_element_by_id("prevButton").click() # prevButton for the previous year
                                  select_period(period,chrome_browser)
                            elif '2021' in period and str(now.year) == '2022':
                                  chrome_browser.find_element_by_id("prevButton").click() # prevButton for the previous year  
                                  select_period(period,chrome_browser)   
                            else:
                                 # This is less likely to happen
                                 chrome_browser.find_element_by_id("nextButton").click() # prevButton for the previous year
                                 select_period(period,chrome_browser)
                                 sys.exit('Nao foi possivel processar o periodo')
                                       
                            time.sleep(3)
                            #indicator_map_file,active_sheet,log_file,browser_webdriver)
                            fill_indicator_elements('RETENTION_ART', ret_on_art_1_3_month_full_path ,active_sheet,log_file,chrome_browser,override)
                            fill_indicator_elements('DSD_MODELS', ret_on_art_dsd_models_full_path , active_sheet,log_file,chrome_browser,override)
                            fill_indicator_elements('12_MONTH_RETENTION', ret_on_art_12_month_full_path,active_sheet,log_file,chrome_browser,override)
                       else:
                             print('Verifique os problemas acima' )
                             log_file.write('Verifique os problemas acima' )
                             sys.exit('Verifique os problemas acima' )  
                  elif form_name == "ATS-Resumo Mensal de Aconselhamento e Testagem em Saude" :
                         #print(workbook.sheetnames)
                         if "Sheet2" in workbook.sheetnames: # ATS Excell template must have 2 sheets  (part A and part B)
                              active_sheet_2 = workbook["Sheet2"]
                              if check_ats_template_integrity(sheet1=active_sheet,sheet2=active_sheet_2,log_file=log_file):
                                   chrome_browser.get(dhis_url)
                                   time.sleep(7)
                                   chrome_browser.find_element_by_id("j_username").send_keys(username)
                                   chrome_browser.find_element_by_id("j_password").send_keys(password)
                                   chrome_browser.find_element_by_id("submit").click()
                                   time.sleep(15)
                                   chrome_browser.get(dhis_url + 'dhis-web-dataentry/index.action')
                                   #tempo para a pagina terminar de carregar
                                   time.sleep(10)
                                   wait = WebDriverWait(chrome_browser, 15)
                                   if user_role == "admin":
                                        xpath="//li[@id='orgUnitj9Inbtfw3Wu']/span/img[contains(@src, '/images/colapse.png')]"
                                        expand_root_tree =wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
                                        expand_root_tree.click()
                                        expand_province_tree('Cidade De Maputo',chrome_browser)
                                        expand_district_tree(district,chrome_browser)
                                        select_province(us_name, chrome_browser)
                                        #time.sleep(4)
                                        select_form(form_name,chrome_browser)
                                        #time.sleep(3)
                                   else:
                                        expand_district_tree(district,chrome_browser)
                                        #time.sleep(4)
                                        select_province(us_name, chrome_browser)
                                        #time.sleep(4)
                                        select_form(form_name,chrome_browser)
                                        #time.sleep(3)
                                   time.sleep(2)
                                   now = datetime.datetime.now()
                                   # Codificacao correcta de caracteres : problema com acentos
                                   if str(now.year) in period:
                                        select_period(period,chrome_browser)
                                   elif '2020' in period and str(now.year) == '2021':
                                        chrome_browser.find_element_by_id("prevButton").click() # prevButton for the previous year
                                        select_period(period,chrome_browser)
                                   elif '2021' in period and str(now.year) == '2022':
                                        chrome_browser.find_element_by_id("prevButton").click() # prevButton for the previous year  
                                        select_period(period,chrome_browser)   
                                   else:
                                        # This is less likely to happen
                                        chrome_browser.find_element_by_id("nextButton").click() # prevButton for the previous year
                                        select_period(period,chrome_browser)
                                        sys.exit('Nao foi possivel processar o periodo')
                                             
                                   time.sleep(3)
                                   fill_indicator_elements('ATS_PART_A_EMERGENCY_WARD', ats_parte_a_emergency_ward_full_path ,active_sheet,log_file,chrome_browser,override)
                                   fill_indicator_elements('ATS_PART_A_OTHER_PITC', ats_parte_a_other_pitc_full_path , active_sheet,log_file,chrome_browser,override)
                                   fill_indicator_elements('ATS_PART_A_INPATIENT_SERVICES', ats_parte_a_inpatient_services_full_path,active_sheet,log_file,chrome_browser,override)
                                   fill_indicator_elements('ATS_PART_A_VCT', ats_parte_a_vct_full_path ,active_sheet,log_file,chrome_browser,override)
                                   fill_indicator_elements('ATS_PART_A_PEDIATRIC_SERVICE', ats_parte_a_pediatric_services_full_path , active_sheet,log_file,chrome_browser,override)
                                   fill_indicator_elements('ATS_PART_B_HTS_INDEX', ats_parte_b_hts_index_full_path,active_sheet_2,log_file,chrome_browser,override)
                              else:
                                   print('Verifique os problemas acima' )
                                   log_file.write('Verifique os problemas acima' )
                                   sys.exit('Verifique os problemas acima' )  
                             
                         else:
                              print('planilha com nome Sheet2 nao foi encontrado'  )
                              log_file.write('planilha com nome Sheet2 nao foi encontrado.\n' )
                              log_file.write("Erro inesperado!! verifique  e corrige os erros  acima.\n")
                              sys.exit("Erro inesperado!! verifique os logs")             
                       
                        
             else:
                 print('planilha com nome %s nao foi encontrado' % sheet_name )
                 log_file.write('planilha com nome %s nao foi encontrado.\n' % sheet_name )
                 log_file.write("Erro inesperado!! verifique  e corrige os erros  acima.\n")
                 sys.exit("Erro inesperado!! verifique os logs")   
               
        else: 
            print('Erro: A unidade sanitaria: %s nao pertence ao distrito  %s no dhis. '  % ( us_name, district ) )
            log_file.write("Erro: A unidade sanitaria: %s nao pertence ao distrito  %s no dhis.\n:"  % ( us_name, district ) )
            log_file.write("Verifique as configuracoes  dhis_config.yaml e tente novamente.\n")
            sys.exit("Erro inesperado!! verifique os logs")
             #ler o ficheiro com dados
      
                        
    except FileNotFoundError as fe:
             print("Erro: Ficheiro nao encontrado ( %s ) "  % excell_location)
             print(fe.args)
             log_file.write("Erro: Ficheiro nao encontrado ( %s )\n"  % excell_location )
             log_file.write("Erro inesperado!! verifique  e corrige os erros  acima\n")
             sys.exit("Erro inesperado!! verifique os logs")   
    except Exception as e:
              print(e.args)
              log_file.write(str(e.args) +'\n')
              log_file.write("Erro inesperado!! verifique  e corrige os erros  acima\n")
              sys.exit("Erro inesperado!! verifique os logs")   

else:
    print("Erro: o ficheiro de parametros dhis_config.yaml nao esta devidamente prenchido, verifique os logs em cima.")
    log_file.write("Verifique as configuracoes  dhis_config.yaml e tente novamente.\n")

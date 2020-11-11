from openpyxl import load_workbook
from helperfunctions import open_config_file

# Open function to open the file "MyFile1.txt"  
# (same directory) in append mode and 
logs_file = open("logs.txt","a+") 

workbook = load_workbook(filename="data/albazine_ct.xlsx")
# grab the active worksheet
ws =workbook['Sheet1']
# Read dhis2 dhis_config
tb_prev_config = open_config_file("mapping/tb_prev.yaml")

for k in range(len(tb_prev_config['TB_PREV_NUMERATOR'])):
    key = str(tb_prev_config['TB_PREV_NUMERATOR'][k].keys())
    f_index = key.find("['")
    s_index = key.find("']")
    indicator = key[f_index+2:s_index]
    #print(indicator)
    xpath = tb_prev_config['TB_PREV_NUMERATOR'][k][indicator]['xpath']
    cell_ref = tb_prev_config['TB_PREV_NUMERATOR'][k][indicator]['cell']
    
    try:
      cell_value = ws[cell_ref].value
      print(cell_ref +" : " + str(cell_value))
      

      
    except:
      print("An exception occurred")
      logs_file.write("An exception occurred in key : %s" % indicator)
      logs_file.close()
    



   
# ws['A2'] = datetime.datetime.now()
# Save the file
# wb.save("sample.xlsx")

